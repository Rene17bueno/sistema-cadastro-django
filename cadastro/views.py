import csv
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
import os
import tempfile
import chardet
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.contrib.auth import login, authenticate, update_session_auth_hash, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.views import LoginView
from django.db import models
from django.utils import timezone
from datetime import datetime
from .models import Cliente, CustomUser
from .forms import ClienteForm, CustomUserCreationForm, CustomUserEditForm, PasswordResetForm, CustomUserProfileForm, CustomPasswordChangeForm
from io import BytesIO
from django.shortcuts import redirect
from django.urls import reverse
from django.contrib.auth import logout
from django.views.decorators.http import require_POST

# Importa módulos necessários para PDF (se reportlab estiver instalado)
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
except ImportError:
    # Se ReportLab não estiver instalado, essas variáveis não existirão.
    # As funções de exportação tratarão a exceção.
    pass


# =============================================
# DECORATORS PARA CONTROLE DE ACESSO
# =============================================
def admin_required(function=None):
    actual_decorator = user_passes_test(
        lambda u: u.is_authenticated and u.tipo_acesso == 'admin',
        login_url='/cadastro/acesso-negado/'
    )
    if function:
        return actual_decorator(function)
    return actual_decorator

def responsavel_ou_admin_required(function=None):
    actual_decorator = user_passes_test(
        lambda u: u.is_authenticated and u.tipo_acesso in ['admin', 'responsavel'],
        login_url='/cadastro/acesso-negado/'
    )
    if function:
        return actual_decorator(function)
    return actual_decorator

def operador_required(function=None):
    actual_decorator = user_passes_test(
        lambda u: u.is_authenticated and u.tipo_acesso in ['admin', 'responsavel', 'operador'],
        login_url='/cadastro/login/'
    )
    if function:
        return actual_decorator(function)
    return actual_decorator

def pode_criar_usuarios_required(function=None):
    actual_decorator = user_passes_test(
        lambda u: u.is_authenticated and (
            u.tipo_acesso == 'admin' or 
            (u.tipo_acesso == 'responsavel' and u.has_perm('cadastro.pode_criar_usuarios'))
        ),
        login_url='/cadastro/acesso-negado/'
    )
    if function:
        return actual_decorator(function)
    return actual_decorator

# =============================================
# VIEWS DE AUTENTICAÇÃO
# =============================================
class CustomLoginView(LoginView):
    template_name = 'cadastro/login.html'
    
    def get_success_url(self):
        return '/cadastro/cadastrar-cliente/'

# VIEW DE LOGOUT
@login_required
def user_logout(request):
    """Desloga o usuário e redireciona para a página de login."""
    logout(request)
    messages.info(request, "Você foi desconectado com sucesso.")
    return redirect('cadastro:login')

# VIEW PARA PÁGINA INICIAL DO APP CADASTRO
@login_required
def home(request):
    return redirect('cadastro:cadastrar_cliente')

# =============================================
# VIEWS DE GERENCIAMENTO DE USUÁRIOS
# =============================================

@login_required
@admin_required
def gerenciar_usuarios(request):
    """Página principal de gerenciamento de usuários"""
    usuarios = CustomUser.objects.all().order_by('-date_joined')
    
    # Estatísticas
    total_usuarios = usuarios.count()
    admins = usuarios.filter(tipo_acesso='admin').count()
    responsaveis = usuarios.filter(tipo_acesso='responsavel').count()
    operadores = usuarios.filter(tipo_acesso='operador').count()
    usuarios_ativos = usuarios.filter(is_active=True).count()
    usuarios_inativos = usuarios.filter(is_active=False).count()
    
    context = {
        'usuarios': usuarios,
        'total_usuarios': total_usuarios,
        'admins': admins,
        'responsaveis': responsaveis,
        'operadores': operadores,
        'usuarios_ativos': usuarios_ativos,
        'usuarios_inativos': usuarios_inativos,
    }
    return render(request, 'cadastro/gerenciar_usuarios.html', context)

@login_required
@admin_required
def listar_usuarios(request):
    """Lista completa de usuários com opções de filtro"""
    usuarios = CustomUser.objects.all().order_by('-date_joined')
    
    # Filtros
    tipo_acesso_filtro = request.GET.get('tipo_acesso', '')
    status_filtro = request.GET.get('status', '')
    
    if tipo_acesso_filtro:
        usuarios = usuarios.filter(tipo_acesso=tipo_acesso_filtro)
    
    if status_filtro:
        if status_filtro == 'ativo':
            usuarios = usuarios.filter(is_active=True)
        elif status_filtro == 'inativo':
            usuarios = usuarios.filter(is_active=False)
    
    context = {
        'usuarios': usuarios,
        'tipo_acesso_filtro': tipo_acesso_filtro,
        'status_filtro': status_filtro,
    }
    return render(request, 'cadastro/listar_usuarios.html', context)

@login_required
def criar_usuario(request):
    # NOTA: Verifique se o decorator pode_criar_usuarios_required deve ser aplicado aqui
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        form.request_user = request.user
        
        if form.is_valid():
            usuario = form.save(commit=False)
            usuario.criado_por = request.user
            usuario.save()
            
            messages.success(request, f'Usuário {usuario.nome_completo} criado com sucesso!')
            return redirect('cadastro:gerenciar_usuarios')
    else:
        form = CustomUserCreationForm()
        if request.user.tipo_acesso == 'responsavel':
            # Responsáveis só podem criar Operadores
            form.fields['tipo_acesso'].choices = [('operador', 'Operador')]
    
    context = {
        'form': form,
        'pode_criar_admin': request.user.tipo_acesso == 'admin'
    }
    return render(request, 'cadastro/criar_usuario.html', context)

@login_required
def editar_usuario(request, usuario_id):
    # NOTA: Verifique se o decorator admin_required ou responsavel_ou_admin_required deve ser aplicado aqui
    usuario = get_object_or_404(CustomUser, id=usuario_id)
    
    if (request.user.tipo_acesso == 'responsavel' and 
        usuario.unidade != request.user.unidade and 
        usuario.criado_por != request.user):
        messages.error(request, 'Você não tem permissão para editar este usuário.')
        return redirect('cadastro:gerenciar_usuarios')
    
    if request.method == 'POST':
        form = CustomUserEditForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, f'Usuário {usuario.nome_completo} atualizado com sucesso!')
            return redirect('cadastro:gerenciar_usuarios')
    else:
        form = CustomUserEditForm(instance=usuario)
        if request.user.tipo_acesso == 'responsavel':
            # Responsáveis só podem editar para Responsável ou Operador
            form.fields['tipo_acesso'].choices = [
                ('responsavel', 'Responsável'),
                ('operador', 'Operador')
            ]
    
    context = {'form': form, 'usuario': usuario}
    return render(request, 'cadastro/editar_usuario.html', context)

@login_required
@admin_required
def ativar_desativar_usuario(request, usuario_id):
    """Ativar/desativar usuário"""
    usuario = get_object_or_404(CustomUser, id=usuario_id)
    
    if request.method == 'POST':
        usuario.is_active = not usuario.is_active
        usuario.save()
        
        acao = "ativado" if usuario.is_active else "desativado"
        messages.success(request, f'Usuário {usuario.nome_completo} {acao} com sucesso!')
    
    return redirect('cadastro:gerenciar_usuarios')

@login_required
@admin_required
def alterar_tipo_acesso(request, usuario_id):
    """Alterar tipo de acesso do usuário"""
    usuario = get_object_or_404(CustomUser, id=usuario_id)
    
    if request.method == 'POST':
        novo_tipo = request.POST.get('tipo_acesso')
        if novo_tipo in ['admin', 'responsavel', 'operador']:
            usuario.tipo_acesso = novo_tipo
            usuario.save()
            messages.success(request, f'Tipo de acesso de {usuario.nome_completo} alterado para {novo_tipo.title()}!')
    
    return redirect('cadastro:gerenciar_usuarios')

@login_required
@admin_required
def excluir_usuario(request, usuario_id):
    """Excluir usuário"""
    usuario = get_object_or_404(CustomUser, id=usuario_id)
    
    if request.method == 'POST':
        if usuario.id == request.user.id:
            messages.error(request, 'Você não pode excluir sua própria conta!')
        else:
            nome_usuario = usuario.nome_completo
            usuario.delete()
            messages.success(request, f'Usuário {nome_usuario} excluído com sucesso!')
    
    return redirect('cadastro:gerenciar_usuarios')

@login_required
def redefinir_senha(request, usuario_id):
    # NOTA: Verifique se o decorator admin_required ou responsavel_ou_admin_required deve ser aplicado aqui
    usuario = get_object_or_404(CustomUser, id=usuario_id)
    
    if (request.user.tipo_acesso == 'responsavel' and 
        usuario.unidade != request.user.unidade and 
        usuario.criado_por != request.user):
        messages.error(request, 'Você não tem permissão para redefinir senha deste usuário.')
        return redirect('cadastro:gerenciar_usuarios')
    
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            nova_senha = form.cleaned_data['nova_senha']
            usuario.set_password(nova_senha)
            usuario.save()
            
            messages.success(request, f'Senha do usuário {usuario.nome_completo} redefinida com sucesso!')
            return redirect('cadastro:gerenciar_usuarios')
    else:
        form = PasswordResetForm()
    
    context = {'form': form, 'usuario': usuario}
    return render(request, 'cadastro/redefinir_senha.html', context)

# =============================================
# VIEWS MEU PERFIL (SEPARADAS)
# =============================================

@login_required
def meu_perfil(request):
    usuario = request.user
    if request.method == 'POST':
        perfil_form = CustomUserProfileForm(request.POST, instance=usuario)
        if perfil_form.is_valid():
            perfil_form.save()
            messages.success(request, 'Perfil atualizado com sucesso!')
            return redirect('cadastro:meu_perfil')
    else:
        perfil_form = CustomUserProfileForm(instance=usuario)
    
    context = {
        'form': perfil_form,
        'user': usuario
    }
    return render(request, 'cadastro/meu_perfil.html', context)

@login_required
def alterar_senha(request):
    if request.method == 'POST':
        senha_form = CustomPasswordChangeForm(request.user, request.POST)
        if senha_form.is_valid():
            user = senha_form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Sua senha foi alterada com sucesso!')
            return redirect('cadastro:meu_perfil')
    else:
        senha_form = CustomPasswordChangeForm(request.user)

    context = {
        'senha_form': senha_form
    }
    return render(request, 'cadastro/alterar_senha.html', context)

@login_required
def acesso_negado(request):
    return render(request, 'cadastro/acesso-negado.html')

# =============================================
# VIEWS EXISTENTES (PROTEGIDAS)
# =============================================

@login_required
@operador_required
def cadastrar_cliente(request):
    unidade_atual = request.POST.get('unidade', '')
    data_atual = request.POST.get('data_cadastro', '')
    
    if not data_atual:
        data_atual = timezone.now().strftime('%Y-%m-%d')
    
    clientes = Cliente.objects.none()
    
    total_clientes = Cliente.objects.count()
    hoje = timezone.now().date()
    clientes_hoje = Cliente.objects.filter(data_cadastro=hoje).count()
    
    if request.method == 'POST':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            form = ClienteForm(request.POST)
            if form.is_valid():
                cliente = form.save()
                return JsonResponse({
                    'success': True,
                    'message': f'Cliente {cliente.codigo_cliente} cadastrado com sucesso!'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                }, status=400)
        
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            messages.success(request, f'Cliente {cliente.codigo_cliente} cadastrado com sucesso!')
            
            unidade_atual = cliente.unidade
            data_atual = cliente.data_cadastro.strftime('%Y-%m-%d')
            
            form = ClienteForm(initial={
                'unidade': unidade_atual,
                'data_cadastro': data_atual
            })
        else:
            messages.error(request, 'Erro ao cadastrar cliente. Verifique os dados.')
    else:
        form = ClienteForm(initial={
            'data_cadastro': timezone.now().strftime('%Y-%m-%d')
        })
    
    return render(request, 'cadastro/cadastro.html', {
        'form': form,
        'clientes': clientes,
        'unidade_atual': unidade_atual,
        'data_atual': data_atual,
        'total_clientes': total_clientes,
        'clientes_hoje': clientes_hoje
    })

@login_required
@responsavel_ou_admin_required
def novos_clientes(request):
    context = {}
    
    if request.method == 'POST' and request.FILES.get('arquivo_csv'):
        arquivo_csv = request.FILES['arquivo_csv']
        
        try:
            resultado = processar_clientes_csv(arquivo_csv)
            
            if resultado:
                with open(resultado['caminho_arquivo'], 'rb') as f:
                    response = HttpResponse(f.read(), content_type='text/plain')
                    response['Content-Disposition'] = f'attachment; filename="{resultado["nome_arquivo"]}"'
                
                messages.success(request, f'Arquivo processado com sucesso! {resultado["registros_processados"]} registros.')
                
                os.remove(resultado['caminho_arquivo'])
                
                return response
            else:
                messages.error(request, 'Erro ao processar o arquivo.')
                
        except Exception as e:
            messages.error(request, f'Erro: {str(e)}')
    
    return render(request, 'cadastro/novos_clientes.html', context)

@login_required
@responsavel_ou_admin_required
def exportar_dados(request):
    unidade_filtro = request.GET.get('unidade', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    formato = request.GET.get('formato', '')
    
    clientes = Cliente.objects.all().order_by('-data_cadastro')
    
    if unidade_filtro:
        clientes = clientes.filter(unidade=unidade_filtro)
    
    if data_inicio:
        try:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            clientes = clientes.filter(data_cadastro__gte=data_inicio_obj)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
            clientes = clientes.filter(data_cadastro__lte=data_fim_obj)
        except ValueError:
            pass
    
    if formato:
        if formato == 'excel':
            return exportar_excel(clientes, unidade_filtro)
        elif formato == 'csv':
            return exportar_csv(clientes, unidade_filtro)
        elif formato == 'pdf':
            return exportar_pdf(clientes, unidade_filtro)
        elif formato == 'txt':
            # Chama a função exportar_txt, que estava faltando
            return exportar_txt(clientes, unidade_filtro)
        else:
            return HttpResponseBadRequest(f"Formato de exportação '{formato}' não suportado.")
    
    context = {
        'unidade_selecionada': unidade_filtro,
        'data_inicio_selecionada': data_inicio,
        'data_fim_selecionada': data_fim,
        'clientes_filtrados': clientes,
        'total_registros': clientes.count(),
        'unidades': ['Maringá', 'Guarapuava', 'Ponta Grossa', 'Norte Pioneiro']
    }
    
    return render(request, 'cadastro/exportar.html', context)

# =============================================
# APIs (PROTEGIDAS)
# =============================================

@login_required
@require_http_methods(["GET"])
def lista_clientes(request):
    unidade_filtro = request.GET.get('unidade', '')
    data_filtro = request.GET.get('data', '')
    
    clientes = Cliente.objects.all().order_by('-id')
    
    if unidade_filtro:
        clientes = clientes.filter(unidade=unidade_filtro)
    
    if data_filtro:
        try:
            data_obj = datetime.strptime(data_filtro, '%Y-%m-%d').date()
            clientes = clientes.filter(data_cadastro=data_obj)
        except ValueError:
            pass
    
    data = []
    for cliente in clientes:
        data.append({
            'id': cliente.id,
            'unidade': cliente.unidade,
            'codigo_cliente': cliente.codigo_cliente,
            'latitude': str(cliente.latitude),
            'longitude': str(cliente.longitude),
            'data_cadastro': cliente.data_cadastro.strftime('%Y-%m-%d'),
        })
    
    return JsonResponse({'clientes': data})

@login_required
@require_http_methods(["GET"])
def detalhe_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    
    data = {
        'id': cliente.id,
        'unidade': cliente.unidade,
        'codigo_cliente': cliente.codigo_cliente,
        'latitude': str(cliente.latitude),
        'longitude': str(cliente.longitude),
        'data_cadastro': cliente.data_cadastro.strftime('%Y-%m-%d'),
    }
    
    return JsonResponse(data)

@csrf_exempt
@login_required
@require_http_methods(["POST"])
def editar_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    
    try:
        data = json.loads(request.body)
        form = ClienteForm(data, instance=cliente)
        
        if form.is_valid():
            cliente = form.save()
            return JsonResponse({
                'success': True,
                'message': f'Cliente {cliente.codigo_cliente} atualizado com sucesso!',
                'cliente': {
                    'id': cliente.id,
                    'unidade': cliente.unidade,
                    'codigo_cliente': cliente.codigo_cliente,
                    'latitude': str(cliente.latitude),
                    'longitude': str(cliente.longitude),
                    'data_cadastro': cliente.data_cadastro.strftime('%Y-%m-%d'),
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            }, status=400)
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
@login_required
@require_http_methods(["DELETE"])
def excluir_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    
    try:
        codigo_cliente = cliente.codigo_cliente
        cliente.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Cliente {codigo_cliente} excluído com sucesso!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
@login_required
@require_http_methods(["POST"])
def validar_cliente(request):
    try:
        data = json.loads(request.body)
        form = ClienteForm(data)
        
        if form.is_valid():
            return JsonResponse({
                'valid': True,
                'message': 'Dados válidos!'
            })
        else:
            return JsonResponse({
                'valid': False,
                'errors': form.errors
            }, status=400)
            
    except Exception as e:
        return JsonResponse({
            'valid': False,
            'error': str(e)
        }, status=500)

@require_POST
def logout_view(request):
    """
    Função para realizar o logout e garantir o redirecionamento.
    O decorador @require_POST força o uso do método POST, evitando o erro 405
    caso alguém tente acessar /logout/ diretamente via GET.
    """
    if request.user.is_authenticated:
        logout(request)
    
    return redirect(reverse('cadastro:login'))

# =============================================
# FUNÇÕES AUXILIARES
# =============================================

def processar_clientes_csv(arquivo_csv):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as temp_file:
            for chunk in arquivo_csv.chunks():
                temp_file.write(chunk)
            temp_path = temp_file.name
        
        with open(temp_path, 'rb') as f:
            raw_data = f.read()
            encoding_result = chardet.detect(raw_data)
            file_encoding = encoding_result['encoding'] or 'latin1'
        
        print(f"Encoding detectado: {file_encoding}")
        
        encodings_to_try = [file_encoding, 'latin1', 'iso-8859-1', 'cp1252', 'utf-8']
        
        geo = None
        for encoding in encodings_to_try:
            try:
                print(f"Tentando ler com encoding: {encoding}")
                geo = pd.read_csv(temp_path, sep=';', encoding=encoding)
                print(f"Sucesso com encoding: {encoding}")
                break
            except UnicodeDecodeError as e:
                print(f"Falha com {encoding}: {e}")
                continue
            except Exception as e:
                print(f"Erro com {encoding}: {e}")
                continue
        
        os.unlink(temp_path)
        
        if geo is None:
            print("Nao foi possível ler o arquivo com nenhum encoding")
            return None
        
        print(f"Colunas encontradas: {list(geo.columns)}")
        if not geo.empty:
            print(f"Primeira linha - Filial: {geo.iloc[0]['Filial']}, Cliente: {geo.iloc[0]['Cliente']}, Coordenadas: {geo.iloc[0]['Coordenadas']}")
        
        filial_mapping = {
            '0001': 'Maringá',
            '0002': 'Guarapuava', 
            '0003': 'Ponta_Grossa',
            '0004': 'Norte_Pioneiro',
            '1': 'Maringá',
            '2': 'Guarapuava',
            '3': 'Ponta_Grossa',
            '4': 'Norte_Pioneiro'
        }
        
        geo['Filial'] = geo['Filial'].astype(str).str.strip()
        geo['Filial_Nome'] = geo['Filial'].map(filial_mapping)
        
        geo['Cliente_Codigo'] = geo['Cliente'].astype(str).str.strip()
        
        resultados = []
        
        for index, row in geo.iterrows():
            try:
                coordenadas = str(row['Coordenadas']).strip()
                cliente_codigo = row['Cliente_Codigo']
                filial_nome = row['Filial_Nome']
                
                data_col = None
                for col in geo.columns:
                    if 'data' in col.lower() or 'inclus' in col.lower():
                        data_col = col
                        break
                
                data_inclusao = row[data_col] if data_col else 'Data não encontrada'
                
                if (not coordenadas or coordenadas == 'nan' or 
                    '000,000000' in coordenadas or coordenadas == '0' or
                    pd.isna(filial_nome) or pd.isna(cliente_codigo)):
                    continue
                
                coord_sem_espacos = coordenadas.replace(' ', '')
                
                partes = coord_sem_espacos.split(',')
                
                if len(partes) >= 4:
                    lat_parte1 = partes[0]
                    lat_parte2 = partes[1]
                    
                    lon_parte1 = partes[2]
                    lon_parte2 = partes[3]
                    
                    latitude = re.sub(r'^(-)0+', r'\1', lat_parte1) + '.' + lat_parte2
                    
                    longitude = re.sub(r'^(-)0+', r'\1', lon_parte1) + '.' + lon_parte2
                    
                    resultados.append({
                        'cliente': cliente_codigo,
                        'latitude': latitude,
                        'longitude': longitude,
                        'filial': filial_nome,
                        'data_inclusao': data_inclusao
                    })
                    print(f"Processado: {cliente_codigo} -> {latitude}, {longitude}")
                    
            except Exception as e:
                print(f"Erro na linha {index}: {e}")
                continue
        
        if not resultados:
            print("Nenhum registro válido encontrado")
            return None
        
        print(f"Total de registros processados: {len(resultados)}")
        
        filial = resultados[0]['filial']
        data_inclusao = resultados[0]['data_inclusao']
        
        try:
            data_obj = datetime.strptime(data_inclusao, '%d/%m/%Y')
            data_formatada = data_obj.strftime('%d-%m-%Y')
        except Exception as e:
            print(f"Erro ao converter data '{data_inclusao}': {e}")
            data_formatada = datetime.now().strftime('%d-%m-%Y')
        
        nome_arquivo = f"{filial}-{data_formatada}.txt"
        
        temp_dir = tempfile.gettempdir()
        caminho_arquivo = os.path.join(temp_dir, nome_arquivo)
        
        with open(caminho_arquivo, 'w', encoding='utf-8') as f:
            for resultado in resultados:
                linha = f"{resultado['cliente']};{resultado['latitude']};{resultado['longitude']}\n"
                f.write(linha)
        
        print(f"Arquivo gerado: {nome_arquivo} com {len(resultados)} registros")
        
        return {
            'caminho_arquivo': caminho_arquivo,
            'nome_arquivo': nome_arquivo,
            'registros_processados': len(resultados)
        }
        
    except Exception as e:
        print(f"Erro geral no processamento: {e}")
        import traceback
        traceback.print_exc()
        return None

# =============================================
# FUNÇÕES DE EXPORTAÇÃO
# =============================================

def exportar_excel(clientes, unidade_filtro):
    if unidade_filtro:
        filename = f"geolocalizacao-{unidade_filtro.lower()}-{timezone.now().strftime('%d-%m-%Y')}.xlsx"
    else:
        filename = f"geolocalizacao-todas-unidades-{timezone.now().strftime('%d-%m-%Y')}.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes Geolocalização"
    
    headers = ['ID', 'Unidade', 'Código Cliente', 'Latitude', 'Longitude', 'Data Cadastro']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row_num, cliente in enumerate(clientes, 2):
        ws.cell(row=row_num, column=1, value=cliente.id)
        ws.cell(row=row_num, column=2, value=cliente.unidade)
        ws.cell(row=row_num, column=3, value=cliente.codigo_cliente)
        ws.cell(row=row_num, column=4, value=str(cliente.latitude))
        ws.cell(row=row_num, column=5, value=str(cliente.longitude))
        ws.cell(row=row_num, column=6, value=cliente.data_cadastro.strftime('%d/%m/%Y'))
    
    column_widths = {
        'A': 8,
        'B': 15,
        'C': 18,
        'D': 20,
        'E': 20,
        'F': 12
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = 'A2'
    
    if len(clientes) > 0:
        ws.auto_filter.ref = f"A1:F{len(clientes) + 1}"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    
    return response

def exportar_csv(clientes, unidade_filtro):
    if unidade_filtro:
        filename = f"{unidade_filtro.lower()}-{timezone.now().strftime('%d-%m-%Y')}.csv"
    else:
        filename = f"todas-unidades-{timezone.now().strftime('%d-%m-%Y')}.csv"
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    response.write('ID,Unidade,Código Cliente,Latitude,Longitude,Data Cadastro\n')
    
    for cliente in clientes:
        linha = f"{cliente.id},{cliente.unidade},{cliente.codigo_cliente},{cliente.latitude},{cliente.longitude},{cliente.data_cadastro.strftime('%d/%m/%Y')}\n"
        response.write(linha)
    
    return response

def exportar_pdf(clientes, unidade_filtro):
    try:
        if unidade_filtro:
            filename = f"geolocalizacao-{unidade_filtro.lower()}-{timezone.now().strftime('%d-%m-%Y')}.pdf"
        else:
            filename = f"geolocalizacao-todas-unidades-{timezone.now().strftime('%d-%m-%Y')}.pdf"
        
        buffer = BytesIO()
        
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            topMargin=0.5*inch,
            bottomMargin=0.5*inch,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1,
            textColor=colors.HexColor('#366092')
        )
        
        if unidade_filtro:
            title_text = f"Relatório de Geolocalização - {unidade_filtro}"
        else:
            title_text = "Relatório de Geolocalização - Todas as Unidades"
            
        title = Paragraph(title_text, title_style)
        elements.append(title)
        
        date_style = ParagraphStyle(
            'CustomDate',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20,
            alignment=1,
        )
        date_text = f"Emitido em: {timezone.now().strftime('%d/%m/%Y às %H:%M')}"
        date_para = Paragraph(date_text, date_style)
        elements.append(date_para)
        
        elements.append(Spacer(1, 0.2*inch))
        
        data = [['Código Cliente', 'Latitude', 'Longitude', 'Unidade', 'Data Cadastro']]
        
        for cliente in clientes:
            data.append([
                str(cliente.codigo_cliente),
                f"{cliente.latitude:.10f}" if cliente.latitude else "N/A",
                f"{cliente.longitude:.10f}" if cliente.longitude else "N/A",
                cliente.unidade,
                cliente.data_cadastro.strftime('%d/%m/%Y')
            ])
        
        # O problema estava aqui, a lista de estilos estava incompleta
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), # Adicionado para completar o estilo
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F5F5F5'), colors.white]) # Estilo zebrado
        ]))
        
        # Adicionar a tabela aos elementos e construir o PDF
        elements.append(table)
        
        def footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 9)
            canvas.drawString(inch, 0.5 * inch, "Página %d" % doc.page)
            canvas.restoreState()

        doc.build(elements, onFirstPage=footer, onLaterPages=footer)
        
        # Resposta HTTP
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        buffer.close()
        return response
        
    except ImportError:
        # Se ReportLab não estiver instalado
        return HttpResponseBadRequest("Biblioteca ReportLab não instalada. Instale com 'pip install reportlab' para exportar em PDF.")
    except Exception as e:
        # Erro genérico
        return HttpResponseBadRequest(f"Erro ao gerar PDF: {str(e)}")


def exportar_txt(clientes, unidade_filtro):
    """
    Exporta dados de clientes no formato TXT (delimitado por ';').
    Inclui APENAS Código Cliente, Latitude e Longitude, SEM cabeçalho.
    """
    import csv # Garante que o módulo csv está disponível
    from django.utils import timezone # Garante que timezone está disponível
    from django.http import HttpResponse # Garante que HttpResponse está disponível

    # Define o nome do arquivo de forma consistente
    if unidade_filtro:
        # Usa o nome da unidade no filename
        filename = f"geolocalizacao-{unidade_filtro.lower()}-{timezone.now().strftime('%d-%m-%Y')}.txt"
    else:
        # Usa "todas-unidades" no filename
        filename = f"geolocalizacao-todas-unidades-{timezone.now().strftime('%d-%m-%Y')}.txt"
    
    # 1. Cria a resposta HTTP com o Content-Type correto para texto
    response = HttpResponse(content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # 2. Cria um objeto escritor CSV, usando ';' como delimitador e '\n' como quebra de linha
    writer = csv.writer(response, delimiter=';', lineterminator='\n')
    
    # 3. Itera sobre o QuerySet e escreve APENAS os 3 campos necessários
    # Não há cabeçalho, conforme solicitado.
    for cliente in clientes:
        writer.writerow([
            cliente.codigo_cliente, 
            # Garante que Latitude e Longitude sejam strings para evitar erros de formatação
            str(cliente.latitude), 
            str(cliente.longitude)
        ])
    
    return response
